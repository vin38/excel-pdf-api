import os, io, traceback
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
import openpyxl.utils
from openpyxl.styles.colors import COLOR_INDEX
from reportlab.lib.pagesizes import A4, A3, letter, landscape, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

# ── Theme color palette (Excel default Office theme) ──────────────────
THEME_COLORS = [
    'FFFFFF','000000','E7E6E6','44546A','4472C4','ED7D31',
    'A9D18E','FF0000','FFFF00','00B0F0','70AD47','FFC000',
    '5B9BD5','FF0000','92D050','00B050','FF7F00','0070C0',
    '7030A0','000000'
]

def apply_tint(hex_color, tint):
    """Apply tint to a hex color (tint: -1 to 1)"""
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        if tint > 0:
            r = int(r + (255 - r) * tint)
            g = int(g + (255 - g) * tint)
            b = int(b + (255 - b) * tint)
        elif tint < 0:
            r = int(r * (1 + tint))
            g = int(g * (1 + tint))
            b = int(b * (1 + tint))
        r = max(0, min(255, r))
        g = max(0, min(255, g))
        b = max(0, min(255, b))
        return '{:02X}{:02X}{:02X}'.format(r, g, b)
    except:
        return hex_color

def resolve_color(color_obj):
    """Resolve openpyxl Color object to hex string"""
    if color_obj is None:
        return None
    try:
        ctype = color_obj.type
        if ctype == 'rgb':
            rgb = color_obj.rgb
            if rgb and len(rgb) >= 6:
                h = rgb[-6:]  # last 6 chars (strip alpha)
                if h.upper() not in ('FFFFFF', '000000', 'FFFFFFF'):
                    return h
        elif ctype == 'theme':
            idx = color_obj.theme
            tint = color_obj.tint or 0
            if idx < len(THEME_COLORS):
                base = THEME_COLORS[idx]
                if tint != 0:
                    base = apply_tint(base, tint)
                if base.upper() not in ('FFFFFF', '000000'):
                    return base
        elif ctype == 'indexed':
            idx = color_obj.indexed
            # Common indexed colors
            indexed_map = {
                2:'FF0000', 3:'00FF00', 4:'0000FF', 5:'FFFF00',
                6:'FF00FF', 7:'00FFFF', 8:'000000', 9:'FFFFFF',
                10:'FF0000', 11:'00FF00', 12:'0000FF', 13:'FFFF00',
                40:'FFFF99', 41:'CCFFFF', 42:'CCFFCC', 43:'FFCC99',
                44:'FF99CC', 45:'CC99FF', 46:'99CCFF', 47:'FF9966',
            }
            if idx in indexed_map:
                return indexed_map[idx]
    except:
        pass
    return None

def hex_to_rl(h):
    if not h or len(h) < 6:
        return None
    try:
        r = int(h[0:2], 16) / 255.0
        g = int(h[2:4], 16) / 255.0
        b = int(h[4:6], 16) / 255.0
        return colors.Color(r, g, b)
    except:
        return None

def get_bg(cell):
    try:
        fill = cell.fill
        if not fill or fill.fill_type in (None, 'none'):
            return None
        fg = fill.fgColor
        h = resolve_color(fg)
        return hex_to_rl(h) if h else None
    except:
        return None

def get_fg(cell):
    try:
        font = cell.font
        if not font or not font.color:
            return None
        h = resolve_color(font.color)
        return hex_to_rl(h) if h else None
    except:
        return None

def get_bold(cell):
    try:
        return bool(cell.font and cell.font.bold)
    except:
        return False

def get_align(cell):
    try:
        if cell.alignment and cell.alignment.horizontal:
            a = cell.alignment.horizontal
            if a in ('right','center','left'):
                return a
        if cell.data_type == 'n' and cell.value is not None:
            return 'right'
    except:
        pass
    return 'left'

def fmt_val(cell):
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, float):
        if v == int(v):
            return '{:,}'.format(int(v))
        return '{:,.2f}'.format(v)
    if isinstance(v, int):
        return '{:,}'.format(v)
    return str(v)

def is_row_empty(row):
    return all(c.value is None for c in row)

@app.route('/', methods=['GET'])
def root():
    return jsonify({'status':'ok','service':'Excel to PDF API v3'})

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status':'ok'})

@app.route('/convert/excel-to-pdf', methods=['POST','OPTIONS'])
def convert():
    if request.method == 'OPTIONS':
        r = jsonify({'ok':True})
        r.headers['Access-Control-Allow-Origin'] = '*'
        r.headers['Access-Control-Allow-Methods'] = 'POST,OPTIONS'
        r.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if 'file' not in request.files:
        return jsonify({'error':'No file'}), 400

    f = request.files['file']
    paper       = request.form.get('paper','A4')
    orient_pref = request.form.get('orientation','auto')
    font_size   = int(request.form.get('font_size', 9))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        paper_map = {'A4':A4,'A3':A3,'Letter':letter}
        base = paper_map.get(paper, A4)
        MARGIN = 36
        all_items = []
        final_page_size = portrait(base)

        for sname in wb.sheetnames:
            ws = wb[sname]

            # Skip completely empty sheets
            if ws.max_row is None or ws.max_column is None:
                continue

            # Find actual data range (skip leading empty rows/cols)
            rows_data = list(ws.iter_rows())
            # Strip leading empty rows
            while rows_data and is_row_empty(rows_data[0]):
                rows_data = rows_data[1:]
            # Strip trailing empty rows
            while rows_data and is_row_empty(rows_data[-1]):
                rows_data = rows_data[:-1]

            if not rows_data:
                continue

            max_col = max(len(r) for r in rows_data)

            # Build data & styles
            data, sgrid = [], []
            for row in rows_data:
                rv, rs = [], []
                for ci in range(max_col):
                    cell = row[ci] if ci < len(row) else None
                    rv.append(fmt_val(cell) if cell else '')
                    rs.append({
                        'bg':    get_bg(cell) if cell else None,
                        'fg':    get_fg(cell) if cell else None,
                        'bold':  get_bold(cell) if cell else False,
                        'align': get_align(cell) if cell else 'left',
                    })
                data.append(rv)
                sgrid.append(rs)

            # Column widths
            col_w = []
            for ci in range(max_col):
                # Get actual column letter from first data row
                actual_ci = rows_data[0][ci].column if ci < len(rows_data[0]) else ci+1
                ltr = openpyxl.utils.get_column_letter(actual_ci)
                d = ws.column_dimensions.get(ltr)
                if d and d.width and d.width > 1:
                    col_w.append(float(d.width) * 5.8)
                else:
                    # Estimate from content
                    mx = max((len(str(data[ri][ci])) for ri in range(len(data))), default=5)
                    col_w.append(max(float(mx) * 6.0, 40.0))

            # Row heights
            row_h = []
            for ri, row in enumerate(rows_data):
                actual_ri = row[0].row if row else ri+1
                d = ws.row_dimensions.get(actual_ri)
                if d and d.height and d.height > 1:
                    row_h.append(float(d.height) * 0.75)
                else:
                    row_h.append(float(font_size) + 5.0)

            total_w = sum(col_w)

            # Orientation
            if orient_pref == 'landscape':
                page_size = landscape(base)
            elif orient_pref == 'portrait':
                page_size = portrait(base)
            else:
                use_p = base[0] - MARGIN*2
                use_l = base[1] - MARGIN*2
                page_size = landscape(base) if total_w > use_p and total_w <= use_l else portrait(base)

            final_page_size = page_size
            usable = page_size[0] - MARGIN*2

            # Scale
            if total_w > usable:
                sc = usable / total_w
                col_w = [w*sc for w in col_w]
                row_h = [max(h*sc, font_size+2) for h in row_h]
                fs = max(6, int(font_size*sc))
            else:
                fs = font_size

            # Build ReportLab table style
            cmds = [
                ('FONTNAME',      (0,0),(-1,-1), 'Helvetica'),
                ('FONTSIZE',      (0,0),(-1,-1), fs),
                ('BACKGROUND',    (0,0),(-1,-1), colors.white),
                ('GRID',          (0,0),(-1,-1), 0.4, colors.Color(0.82,0.82,0.82)),
                ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
                ('LEFTPADDING',   (0,0),(-1,-1), 3),
                ('RIGHTPADDING',  (0,0),(-1,-1), 3),
                ('TOPPADDING',    (0,0),(-1,-1), 2),
                ('BOTTOMPADDING', (0,0),(-1,-1), 2),
            ]

            rl_align = {'right':'RIGHT','center':'CENTER','left':'LEFT'}

            for ri, rs in enumerate(sgrid):
                for ci, st in enumerate(rs):
                    coord = (ci, ri)
                    if st['bg']:
                        cmds.append(('BACKGROUND', coord, coord, st['bg']))
                    if st['fg']:
                        cmds.append(('TEXTCOLOR',  coord, coord, st['fg']))
                    if st['bold']:
                        cmds.append(('FONTNAME',   coord, coord, 'Helvetica-Bold'))
                    cmds.append(('ALIGN', coord, coord, rl_align.get(st['align'],'LEFT')))

            tbl = Table(data, colWidths=col_w, rowHeights=row_h, repeatRows=1)
            tbl.setStyle(TableStyle(cmds))
            all_items.append(tbl)

        if not all_items:
            return jsonify({'error':'No data found'}), 400

        pdf_buf = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buf, pagesize=final_page_size,
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN,  bottomMargin=MARGIN+12,
        )
        doc.build(all_items)
        pdf_buf.seek(0)

        out = (f.filename or 'output').rsplit('.',1)[0] + '.pdf'
        return send_file(pdf_buf, mimetype='application/pdf',
                         as_attachment=True, download_name=out)

    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
